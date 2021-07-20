from tkinter import *
from openpyxl import *
import sqlite3

#/home/magni/github/zavrsni_rad/izvjestaj1.xlsx

def ucitavanje_automatsko(event):
    izvjestaj_putanja = putanja.get()
    wb = load_workbook(izvjestaj_putanja)
    ws = wb.active

    djelatnik = ws.cell(row=5, column=3).value
    djelatnikLabel = Label(lf1, text=f"Djelatnik: {djelatnik}").grid(row=1, column=0)
    
    nazivRacunala = ws.cell(row=6, column=3).value
    nazivRacunalaLabel = Label(lf1, text=f"Naziv Računala: {nazivRacunala}").grid(row=2, column=0)

    korisnickoIme = ws.cell(row=7, column=3).value
    korisnickoImeLabel = Label(lf1, text=f"Korisničko ime: {korisnickoIme}").grid(row=3, column=0)

    lozinka = ws.cell(row=8, column=3).value
    lozinkaLabel = Label(lf1, text=f"Lozinka: {lozinka}").grid(row=4, column=0)

    inventurniBroj = ws.cell(row=9, column=3).value
    inventurniBrojLabel = Label(lf1, text=f"Inventurni broj: {inventurniBroj}").grid(row=5, column=0)

    macAdresa = ws.cell(row=10, column=3).value
    macAdresaLabel = Label(lf1, text=f"MAC adresa: {macAdresa}").grid(row=6, column=0)

    operativniSustav = ws.cell(row=14, column=3).value
    operativniSustavLabel = Label(lf1, text=f"Operativni sustav: {operativniSustav}").grid(row=7, column=0)

    dodatniProgram1 = ws.cell(row=16, column=2).value
    dodatniProgram1Label = Label(lf1, text=f"Dodatni program: {dodatniProgram1}").grid(row=8, column=0)

    dodatniProgram2 = ws.cell(row=17, column=2).value
    dodatniProgram2Label = Label(lf1, text=f"Dodatni program: {dodatniProgram2}").grid(row=9, column=0)

    dodatniProgram3 = ws.cell(row=18, column=2).value
    dodatniProgram3Label = Label(lf1, text=f"Dodatni program: {dodatniProgram3}").grid(row=10, column=0)

    dodatniProgram4 = ws.cell(row=19, column=2).value
    dodatniProgram4Label = Label(lf1, text=f"Dodatni program: {dodatniProgram4}").grid(row=11, column=0)

    dodatniProgram5 = ws.cell(row=20, column=2).value
    dodatniProgram5Label = Label(lf1, text=f"Dodatni program: {dodatniProgram5}").grid(row=12, column=0)

    dodatniProgram6 = ws.cell(row=21, column=2).value
    dodatniProgram6Label = Label(lf1, text=f"Dodatni program: {dodatniProgram6}").grid(row=13, column=0)

    dodatniProgram7 = ws.cell(row=22, column=2).value
    dodatniProgram7Label = Label(lf1, text=f"Dodatni program: {dodatniProgram7}").grid(row=14, column=0)

    dodatniProgram8 = ws.cell(row=23, column=2).value
    dodatniProgram8Label = Label(lf1, text=f"Dodatni program: {dodatniProgram8}").grid(row=15, column=0)

    dodatniProgram9 = ws.cell(row=24, column=2).value
    dodatniProgram9Label = Label(lf1, text=f"Dodatni program: {dodatniProgram9}").grid(row=16, column=0)

    dodatniProgram10 = ws.cell(row=25, column=2).value
    dodatniProgram10Label = Label(lf1, text=f"Dodatni program: {dodatniProgram10}").grid(row=17, column=0)

    dodatniProgram11 = ws.cell(row=26, column=2).value
    dodatniProgram11Label = Label(lf1, text=f"Dodatni program: {dodatniProgram11}").grid(row=18, column=0)

    dodatniProgram12 = ws.cell(row=27, column=2).value
    dodatniProgram12Label = Label(lf1, text=f"Dodatni program: {dodatniProgram12}").grid(row=19, column=0)

    dodatniProgram13 = ws.cell(row=28, column=2).value
    dodatniProgram13Label = Label(lf1, text=f"Dodatni program: {dodatniProgram13}").grid(row=20, column=0)

    dodatniProgram14 = ws.cell(row=29, column=2).value
    dodatniProgram14Label = Label(lf1, text=f"Dodatni program: {dodatniProgram14}").grid(row=21, column=0)

    dodatniProgram15 = ws.cell(row=30, column=2).value
    dodatniProgram15Label = Label(lf1, text=f"Dodatni program: {dodatniProgram15}").grid(row=22, column=0)

    model = ws.cell(row=5, column=6).value
    modelLabel = Label(rf1, text=f"Model: {model}").grid(row=1, column=0)

    cpu = ws.cell(row=6, column=6).value
    cpuLabel = Label(rf1, text=f"Cpu: {cpu}").grid(row=2, column=0)

    ram = ws.cell(row=7, column=6).value
    ramLabel = Label(rf1, text=f"RAM: {ram}").grid(row=3, column=0)

    mreznaKartica = ws.cell(row=8, column=6).value
    mreznaKarticaLabel = Label(rf1, text=f"Mrežna kartica: {mreznaKartica}").grid(row=4, column=0)

    grafickaKartica = ws.cell(row=9, column=6).value
    grafickaKarticaLabel = Label(rf1, text=f"Grafička kartica: {grafickaKartica}").grid(row=5, column=0)

    glavnaParticija = ws.cell(row=10, column=6).value
    glavnaParticijaLabel = Label(rf1, text=f"Glavna particija: {glavnaParticija}").grid(row=6, column=0)

    dodatneParticije = ws.cell(row=11, column=6).value
    dodatneParticijeLabel = Label(rf1, text=f"Dodatne particije: {dodatneParticije}").grid(row=7, column=0)

    napomena1 = ws.cell(row=14, column=5).value
    napomena1Label = Label(rf1, text=f"Napomena: {napomena1}").grid(row=8, column=0)

    napomena2 = ws.cell(row=15, column=5).value
    napomena2Label = Label(rf1, text=f"Napomena: {napomena2}").grid(row=9, column=0)

    napomena3 = ws.cell(row=16, column=5).value
    napomena3Label = Label(rf1, text=f"Napomena: {napomena3}").grid(row=10, column=0)

    napomena4 = ws.cell(row=17, column=5).value
    napomena4Label = Label(rf1, text=f"Napomena: {napomena4}").grid(row=11, column=0)

    napomena5 = ws.cell(row=18, column=5).value
    napomena5Label = Label(rf1, text=f"Napomena: {napomena5}").grid(row=12, column=0)

    napomena6 = ws.cell(row=19, column=5).value
    napomena6Label = Label(rf1, text=f"Napomena: {napomena6}").grid(row=13, column=0)

    napomena7 = ws.cell(row=20, column=5).value
    napomena7Label = Label(rf1, text=f"Napomena: {napomena7}").grid(row=14, column=0)

    napomena8 = ws.cell(row=21, column=5).value
    napomena8Label = Label(rf1, text=f"Napomena: {napomena8}").grid(row=15, column=0)

    napomena9 = ws.cell(row=22, column=5).value
    napomena9Label = Label(rf1, text=f"Napomena: {napomena9}").grid(row=16, column=0)

    napomena10 = ws.cell(row=23, column=5).value
    napomena10Label = Label(rf1, text=f"Napomena: {napomena10}").grid(row=17, column=0)

    napomena11 = ws.cell(row=24, column=5).value
    napomena11Label = Label(rf1, text=f"Napomena: {napomena11}").grid(row=18, column=0)

    napomena12 = ws.cell(row=25, column=5).value
    napomena12Label = Label(rf1, text=f"Napomena: {napomena12}").grid(row=19, column=0)

    napomena13 = ws.cell(row=26, column=5).value
    napomena13Label = Label(rf1, text=f"Napomena: {napomena13}").grid(row=20, column=0)

    napomena14 = ws.cell(row=27, column=5).value
    napomena14Label = Label(rf1, text=f"Napomena: {napomena14}").grid(row=21, column=0)

    napomena15 = ws.cell(row=28, column=5).value
    napomena15Label = Label(rf1, text=f"Napomena: {napomena15}").grid(row=22, column=0)

    napomena16 = ws.cell(row=29, column=5).value
    napomena16Label = Label(rf1, text=f"Napomena: {napomena16}").grid(row=23, column=0)

    napomena17 = ws.cell(row=30, column=5).value
    napomena17Label = Label(rf1, text=f"Napomena: {napomena17}").grid(row=24, column=0)

    datum = ws.cell(row=32, column=3).value
    datumlabel = Label(lf2, text=f"Datum: {datum}").grid(row=0, column=0)

    itDjelatnik = ws.cell(row=32, column=6).value
    itDjelatnikLabel = Label(rf2, text=f"IT djelatnik: {itDjelatnik}").grid(row=0, column=0)

    conn = sqlite3.connect('izvjestajiBaza.db')
    c = conn.cursor()

    def create_table():
        c.execute("CREATE TABLE IF NOT EXISTS izvjestajTablica(inventurniBroj INTEGER, djelatnik TEXT, "
                  "nazivRacunala TEXT, korisnickoIme TEXT, lozinka TEXT, macAdresa TEXT, operativniSustav TEXT, "
                  "dodatniProgram1 TEXT, dodatniProgram2 TEXT, dodatniProgram3 TEXT, dodatniProgram4 TEXT, "
                  "dodatniProgram5 TEXT, dodatniProgram6 TEXT, dodatniProgram7 TEXT, dodatniProgram8 TEXT, "
                  "dodatniProgram9 TEXT, dodatniProgram10 TEXT, dodatniProgram11 TEXT, dodatniProgram12 TEXT, "
                  "dodatniProgram13 TEXT, dodatniProgram14 TEXT, dodatniProgram15 TEXT, model TEXT, cpu TEXT, "
                  "ram TEXT, mreznaKartica TEXT, grafickaKartica TEXT, glavnaParticija TEXT, dodatneParticije TEXT, "
                  "napomena1 TEXT, napomena2 TEXT, napomena3 TEXT, napomena4 TEXT, napomena5 TEXT, napomena6 TEXT, "
                  "napomena7 TEXT, napomena8 TEXT, napomena9 TEXT, napomena10 TEXT, napomena11 TEXT, napomena12 TEXT, "
                  "napomena13 TEXT, napomena14 TEXT, napomena15 TEXT, napomena16 TEXT, napomena17 TEXT, datum TEXT, "
                  "itDjelatnik TEXT)")

    def data_entry():
        c.execute("INSERT INTO izvjestajTablica(inventurniBroj, djelatnik, nazivRacunala, korisnickoIme, lozinka, "
                  "macAdresa, operativniSustav, dodatniProgram1, dodatniProgram2, dodatniProgram3, dodatniProgram4, "
                  "dodatniProgram5, dodatniProgram6, dodatniProgram7, dodatniProgram8, dodatniProgram9, "
                  "dodatniProgram10, dodatniProgram11, dodatniProgram12, dodatniProgram13, dodatniProgram14, "
                  "dodatniProgram15, model, cpu, ram, mreznaKartica, grafickaKartica, glavnaParticija, "
                  "dodatneParticije, napomena1, napomena2, napomena3, napomena4, napomena5, napomena6, napomena7, "
                  "napomena8, napomena9, napomena10, napomena11, napomena12, napomena13, napomena14, napomena15, "
                  "napomena16, napomena17, datum, itDjelatnik) VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, "
                  "?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, "
                  "?)", (inventurniBroj, djelatnik, nazivRacunala, korisnickoIme, lozinka, macAdresa,
                        operativniSustav, dodatniProgram1, dodatniProgram2, dodatniProgram3, dodatniProgram4,
                        dodatniProgram5, dodatniProgram6, dodatniProgram7, dodatniProgram8, dodatniProgram9,
                        dodatniProgram10, dodatniProgram11, dodatniProgram12, dodatniProgram13,
                        dodatniProgram14, dodatniProgram15, model, cpu, ram, mreznaKartica,
                        grafickaKartica, glavnaParticija, dodatneParticije, napomena1, napomena2, napomena3,
                        napomena4, napomena5, napomena6, napomena7, napomena8, napomena9, napomena10,
                        napomena11, napomena12, napomena13, napomena14, napomena15, napomena16,
                        napomena17, datum, itDjelatnik))
        conn.commit()


    def run_all(event):
        create_table()
        data_entry()
        gotovoLabel = Label(lf3, text="Pohranjeno!").grid(row=1, column=0)
        



    pohraniPodatkeButton = Button(lf3, width=15, height=1)
    pohraniPodatkeButton.config(text='Pohrani u BP!', fg='red', bg='silver')
    pohraniPodatkeButton.bind("<Button-1>", run_all)
    pohraniPodatkeButton.bind("<Button-2>", run_all)
    pohraniPodatkeButton.bind("<Button-3>", run_all)
    pohraniPodatkeButton.grid(row=0, column=0)
    
    
    
root = Tk()
root.geometry("1200x800")
root.wm_title('Unesite putanju do izvještaja, sa ekstenzijom:')
frame = Frame(root)
frame.grid()

putanja = StringVar()

lf1 = Frame(frame)
lf1.grid(row=1, column=0)

lf2 = Frame(frame)
lf2.grid(row=2, column=0)

lf3 = Frame(frame)
lf3.grid(row=3, column=0)

rf1 = Frame(frame)
rf1.grid(row=1, column=2)

rf2 = Frame(frame)
rf2.grid(row=2, column=2)

rf3 = Frame(frame)
rf3.grid(row=3, column=2)

lokacijaEntry = Entry(frame, textvariable=putanja, width=50).grid(row=0, column=1)

ucitajIzvjestajButton = Button(frame, width=15, height=1, text='Učitaj izvještaj', fg='red', bg='silver')
ucitajIzvjestajButton.bind("<Button-1>", ucitavanje_automatsko)
ucitajIzvjestajButton.bind("<Button-2>", ucitavanje_automatsko)
ucitajIzvjestajButton.bind("<Button-3>", ucitavanje_automatsko)
ucitajIzvjestajButton.grid(row=0, column=0)

root.mainloop()